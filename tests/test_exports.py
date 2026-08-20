import csv
import json

from openpyxl import load_workbook

from acquisition.session_recorder import SessionRecorder
from app.models import FilterSettings, Participant, SessionConfig, SessionMetadata
from storage.paths import DataPaths
from storage.session_repository import SessionRepository


def _recording(sample_factory):
    metadata = SessionMetadata(
        Participant(participant_id="EXPORT", identifier_only=True),
        SessionConfig(duration_s=1, filter_settings=FilterSettings(kind="none")),
        "simulator",
        synthetic_data=True,
    )
    recorder = SessionRecorder(metadata)
    for index in range(11):
        recorder.add_sample(sample_factory(index / 10, x=index / 100, sequence=index))
    recorder.finalize()
    return recorder


def test_complete_export_set(tmp_path, sample_factory):
    recorder = _recording(sample_factory)
    repo = SessionRepository(DataPaths(tmp_path / "data"))
    artifacts = repo.save(recorder.metadata, recorder.samples, recorder.filtered_samples)
    assert artifacts.csv_path.read_bytes().startswith(b"\xef\xbb\xbf")
    with artifacts.csv_path.open(encoding="utf-8-sig") as handle:
        header = next(csv.reader(handle, delimiter=";"))
    assert "cop_x" in header
    assert {"raw_total_weight_kg", "raw_cop_x", "raw_cop_y"}.issubset(header)
    workbook = load_workbook(artifacts.xlsx_path, read_only=True)
    assert workbook.sheetnames == [
        "Metadane",
        "Dane surowe",
        "Dane przetworzone",
        "Wyniki",
        "Jakość",
    ]
    workbook.close()
    metadata = json.loads(artifacts.json_path.read_text(encoding="utf-8"))
    assert metadata["synthetic_data"] is True
    assert "nie stanowią diagnozy medycznej" in artifacts.markdown_path.read_text(encoding="utf-8")


def test_list_sessions_filters_participant(tmp_path, sample_factory):
    recorder = _recording(sample_factory)
    repo = SessionRepository(DataPaths(tmp_path / "data"))
    repo.save(recorder.metadata, recorder.samples, recorder.filtered_samples)
    assert len(repo.list_sessions("EXPORT")) == 1
    assert repo.list_sessions("OTHER") == []


def test_load_stabilogram_from_saved_session(tmp_path, sample_factory):
    recorder = _recording(sample_factory)
    repo = SessionRepository(DataPaths(tmp_path / "data"))
    repo.save(recorder.metadata, recorder.samples, recorder.filtered_samples)
    record = repo.list_sessions("EXPORT")[0]
    points = repo.load_stabilogram(record)
    assert len(points) == 11
    assert points[-1][0] == 0.1
