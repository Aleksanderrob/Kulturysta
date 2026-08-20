"""Multi-sheet XLSX export."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.models import BoardSample, SessionMetadata


def _flat_rows(mapping: dict[str, Any], prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    for key, value in mapping.items():
        name = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            rows.extend(_flat_rows(value, name))
        elif isinstance(value, (list, tuple)):
            rows.append((name, ", ".join(map(str, value))))
        else:
            rows.append((name, value))
    return rows


def _write_samples(sheet: Any, samples: list[BoardSample], processed: bool) -> None:
    columns = [
        "timestamp_monotonic",
        "timestamp_wall",
        "sequence_number",
        "top_left",
        "top_right",
        "bottom_left",
        "bottom_right",
        "total_weight_kg",
        "raw_total_weight_kg",
        "cop_x",
        "cop_y",
        "raw_cop_x",
        "raw_cop_y",
        "filtered_cop_x",
        "filtered_cop_y",
        "cop_unit",
        "connection_ok",
        "quality_flags",
    ]
    sheet.append(columns)
    for sample in samples:
        row = sample.to_dict()
        if processed:
            row["cop_x"] = row.get("filtered_cop_x")
            row["cop_y"] = row.get("filtered_cop_y")
        else:
            row["total_weight_kg"] = (
                row.get("raw_total_weight_kg")
                if row.get("raw_total_weight_kg") is not None
                else row.get("total_weight_kg")
            )
            row["cop_x"] = (
                row.get("raw_cop_x") if row.get("raw_cop_x") is not None else row.get("cop_x")
            )
            row["cop_y"] = (
                row.get("raw_cop_y") if row.get("raw_cop_y") is not None else row.get("cop_y")
            )
        row["quality_flags"] = "|".join(row["quality_flags"])
        sheet.append([row.get(key) for key in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def export_excel(
    metadata: SessionMetadata,
    raw_samples: list[BoardSample],
    processed_samples: list[BoardSample],
    path: Path,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    metadata_sheet = workbook.active
    metadata_sheet.title = "Metadane"
    metadata_sheet.append(["Pole", "Wartość"])
    for row in _flat_rows(metadata.to_dict()):
        metadata_sheet.append(row)
    raw_sheet = workbook.create_sheet("Dane surowe")
    _write_samples(raw_sheet, raw_samples, False)
    processed_sheet = workbook.create_sheet("Dane przetworzone")
    _write_samples(processed_sheet, processed_samples, True)
    results_sheet = workbook.create_sheet("Wyniki")
    results_sheet.append(["Parametr", "Wartość"])
    for key, value in metadata.metrics.items():
        results_sheet.append([key, value])
    quality_sheet = workbook.create_sheet("Jakość")
    quality_sheet.append(["Klasyfikacja", metadata.quality_rating.value])
    quality_sheet.append(["Flagi"])
    for flag in metadata.quality_flags:
        quality_sheet.append([flag.value])
    workbook.save(path)
    # Opening the newly written workbook catches common serialization/corruption problems.
    load_workbook(path, read_only=True).close()
    return path
